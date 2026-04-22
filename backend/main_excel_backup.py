from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional, List, Dict, Any
import pandas as pd
import json
from pathlib import Path
from datetime import datetime
import os
from openpyxl import load_workbook, Workbook

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Data paths
DATA_DIR = Path(__file__).parent.parent / "data"
DATA_DIR.mkdir(exist_ok=True)

EXCEL_FILE = DATA_DIR / "graftcare_new.xlsx"

# Models
class VendorModel(BaseModel):
    id: Optional[str] = None
    name: str
    contact_person: str
    phone: str
    gstin: str
    dl1: str
    dl2: str
    address: str
    city: str
    pincode: str
    bank_name: str
    bank_acc: str
    bank_ifsc: str

class ProductModel(BaseModel):
    id: Optional[str] = None
    name: str
    hsn_code: str
    pack: str = ""
    company: str = ""
    scheme: str = ""
    cost_price: float
    mrp: float
    gst_rate: float

class PurchaseInvoiceModel(BaseModel):
    id: Optional[str] = None
    # Vendor details
    vendor_name: str  # Mandatory
    contact_person: Optional[str] = None
    vendor_gstin: str  # Mandatory
    dl_no_1: Optional[str] = None
    dl_no_2: Optional[str] = None
    phone: str  # Mandatory
    address: Optional[str] = None
    city: str  # Mandatory
    pincode: str  # Mandatory
    # Bank details (optional)
    vendor_bank_name: Optional[str] = None
    vendor_account_no: Optional[str] = None
    vendor_ifsc: Optional[str] = None
    # Invoice header
    vendor_invoice_no: str  # Mandatory
    invoice_date: str  # Mandatory
    note: Optional[str] = None
    # Payment details
    payment_mode: str  # Mandatory
    amount_paid: float  # Mandatory
    paid_by: str  # Mandatory
    # Products (stored as JSON array)
    products: list  # Array of {product_id, name, qty, batch, expiry, mrp, buy_rate, free, discount, gst}
    # Summary
    subtotal: float
    total_gst: float
    discount_amount: float
    grand_total: float
    # Tracking
    ordered_products: Optional[list] = None
    delivered_products: Optional[list] = None
    expected_delivery: Optional[list] = None
    created_date: Optional[str] = None
    updated_date: Optional[str] = None

class PurchaseModel(BaseModel):  # Keep for backward compatibility
    id: Optional[str] = None
    vendor_id: str | int
    product_id: str | int
    quantity: int
    batch_no: str
    expiry_date: str
    invoice_no: str
    invoice_date: str
    amount_paid: float
    paid_by: str

class SalesModel(BaseModel):
    id: Optional[str] = None
    customer_name: str
    customer_phone: str
    product_id: str
    quantity: int
    selling_price: float
    sale_date: str
    status: str = "draft"  # draft or final

class DraftInvoiceModel(BaseModel):
    id: Optional[str] = None
    data: dict  # Complete invoice data (all proforma fields)
    created_date: Optional[str] = None
    updated_date: Optional[str] = None

class InvoiceModel(BaseModel):
    id: Optional[str] = None
    draft_id: Optional[str] = None
    data: dict  # Complete invoice data
    finalized_date: Optional[str] = None

class FinalizeModel(BaseModel):
    invNo: Optional[str] = None
    data: Optional[dict] = None

class StockAddModel(BaseModel):
    product_id: str
    product_name: str
    quantity: int
    batch_no: str
    expiry_date: str
    vendor_id: Optional[str] = None
    vendor_name: Optional[str] = None
    invoice_no: Optional[str] = None
    invoice_date: Optional[str] = None
    cost_price: Optional[float] = None
    amount_paid: Optional[float] = None
    paid_by: Optional[str] = None

def gen_id():
    return datetime.now().strftime("%Y%m%d%H%M%S") + str(abs(hash(os.urandom(4))) % 10000)

def init_excel():
    """Create or update Excel file with required sheets"""
    if EXCEL_FILE.exists():
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    # Ensure Vendors sheet exists
    if "Vendors" not in wb.sheetnames:
        ws = wb.create_sheet("Vendors")
        ws.append(['id','name','contact_person','phone','gstin','dl1','dl2','address','city','pincode','bank_name','bank_acc','bank_ifsc'])

    # Ensure Products sheet exists
    if "Products" not in wb.sheetnames:
        ws = wb.create_sheet("Products")
        ws.append(['id','name','hsn_code','pack','company','scheme','cost_price','mrp','gst_rate'])

    # Ensure Purchases sheet exists
    if "Purchases" not in wb.sheetnames:
        ws = wb.create_sheet("Purchases")
        ws.append(['id','vendor_id','product_id','quantity','batch_no','expiry_date','invoice_no','invoice_date','amount_paid','paid_by'])

    # Ensure Sales sheet exists with status column
    if "Sales" not in wb.sheetnames:
        ws = wb.create_sheet("Sales")
        ws.append(['id','customer_name','customer_phone','product_id','quantity','selling_price','sale_date','status'])
    else:
        # Add status column if it doesn't exist
        ws = wb["Sales"]
        if 'status' not in ws[1]:
            ws['I1'] = 'status'

    # Ensure Drafts sheet exists (in-progress proforma invoices)
    if "Drafts" not in wb.sheetnames:
        ws = wb.create_sheet("Drafts")
        ws.append(['id','data','created_date','updated_date'])

    # Ensure Proformas sheet exists (finalized proforma invoices)
    if "Proformas" not in wb.sheetnames:
        ws = wb.create_sheet("Proformas")
        ws.append(['id','data','created_date','updated_date'])

    # Ensure Invoices sheet exists (finalized sales invoices from proformas)
    if "Invoices" not in wb.sheetnames:
        ws = wb.create_sheet("Invoices")
        ws.append(['id','proforma_id','data','finalized_date'])

    # Ensure PurchaseInvoices sheet exists
    if "PurchaseInvoices" not in wb.sheetnames:
        ws = wb.create_sheet("PurchaseInvoices")
        ws.append(['id','vendor_name','vendor_gstin','phone','city','pincode','vendor_invoice_no','invoice_date','payment_mode','amount_paid','paid_by','products','subtotal','total_gst','discount_amount','grand_total','created_date','updated_date'])

    # Ensure Stock sheet exists (inventory tracking)
    if "Stock" not in wb.sheetnames:
        ws = wb.create_sheet("Stock")
        ws.append(['product_id','product_name','Purchased','Sold','Available'])

    # Ensure Counters sheet exists (for sequential invoice numbers)
    if "Counters" not in wb.sheetnames:
        ws = wb.create_sheet("Counters")
        ws.append(['counter_name','current_value'])
        ws.append(['retail_invoice',0])
        ws.append(['patient_invoice',0])
        ws.append(['purchase_invoice',0])

    # Ensure Expiry sheet exists (expiry tracking)
    if "Expiry" not in wb.sheetnames:
        ws = wb.create_sheet("Expiry")
        ws.append(['product_id','product_name','batch','expiry_date','quantity','purchase_id'])

    # Ensure States sheet exists (state-wise SGST/CGST rates)
    if "States" not in wb.sheetnames:
        ws = wb.create_sheet("States")
        ws.append(['state_code','state_name','sgst_rate','cgst_rate','igst_rate'])
        # Add all Indian states with standard GST rates
        states = [
            ['AP', 'Andhra Pradesh', 9, 9, 18],
            ['AR', 'Arunachal Pradesh', 9, 9, 18],
            ['AS', 'Assam', 9, 9, 18],
            ['BR', 'Bihar', 9, 9, 18],
            ['CG', 'Chhattisgarh', 9, 9, 18],
            ['CT', 'Chandigarh', 9, 9, 18],
            ['DD', 'Daman and Diu', 9, 9, 18],
            ['DL', 'Delhi', 9, 9, 18],
            ['DN', 'Dadra and Nagar Haveli', 9, 9, 18],
            ['GA', 'Goa', 9, 9, 18],
            ['GJ', 'Gujarat', 9, 9, 18],
            ['HR', 'Haryana', 9, 9, 18],
            ['HP', 'Himachal Pradesh', 9, 9, 18],
            ['JK', 'Jammu and Kashmir', 9, 9, 18],
            ['JH', 'Jharkhand', 9, 9, 18],
            ['KA', 'Karnataka', 9, 9, 18],
            ['KL', 'Kerala', 9, 9, 18],
            ['LA', 'Ladakh', 9, 9, 18],
            ['LD', 'Lakshadweep', 9, 9, 18],
            ['MP', 'Madhya Pradesh', 9, 9, 18],
            ['MH', 'Maharashtra', 9, 9, 18],
            ['MN', 'Manipur', 9, 9, 18],
            ['ML', 'Meghalaya', 9, 9, 18],
            ['MZ', 'Mizoram', 9, 9, 18],
            ['NL', 'Nagaland', 9, 9, 18],
            ['OR', 'Odisha', 9, 9, 18],
            ['PB', 'Punjab', 9, 9, 18],
            ['RJ', 'Rajasthan', 9, 9, 18],
            ['SK', 'Sikkim', 9, 9, 18],
            ['TN', 'Tamil Nadu', 9, 9, 18],
            ['TR', 'Tripura', 9, 9, 18],
            ['TS', 'Telangana', 9, 9, 18],
            ['UP', 'Uttar Pradesh', 9, 9, 18],
            ['UK', 'Uttarakhand', 9, 9, 18],
            ['WB', 'West Bengal', 9, 9, 18]
        ]
        for state in states:
            ws.append(state)

    wb.save(EXCEL_FILE)

init_excel()

def read_sheet(sheet_name):
    """Read Excel sheet"""
    try:
        # Force 'id' and 'data' columns to be read as strings
        dtype_dict = {}
        if sheet_name in ['Vendors', 'Products', 'Drafts', 'Proformas', 'Invoices', 'Purchases']:
            dtype_dict['id'] = str
        if sheet_name in ['Drafts', 'Proformas', 'Invoices']:
            dtype_dict['data'] = str  # Keep data as string for JSON parsing

        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name, dtype=dtype_dict if dtype_dict else None)
        df = df.fillna('')
        return df
    except:
        return pd.DataFrame()

def write_sheet(df, sheet_name):
    """Write to Excel sheet"""
    df = df.copy()

    # Ensure ID columns are stored as strings to prevent precision loss
    if 'id' in df.columns:
        df['id'] = df['id'].astype(str)

    # Ensure data column stays as JSON string (don't let pandas convert it)
    if 'data' in df.columns:
        df['data'] = df['data'].astype(str)

    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

def deduct_stock_for_items(items_list):
    """Robust stock deduction function for invoice items"""
    if not items_list:
        return True

    try:
        stock_df = read_sheet('Stock')
        products_df = read_sheet('Products')

        # Ensure Stock sheet structure
        expected_columns = ['product_id', 'product_name', 'Purchased', 'Sold', 'Available']
        if stock_df.empty:
            stock_df = pd.DataFrame(columns=expected_columns)

        changes_made = False

        for item in items_list:
            if not item or item.get('qty', 0) <= 0:
                continue

            product_name = item.get('name', item.get('productName', ''))
            qty = int(item.get('qty', 0))

            if not product_name:
                print(f"Warning: Item has no name, skipping")
                continue

            print(f"\n→ Deducting stock for: {product_name} (Qty: {qty})")

            # Find product ID
            product_db_id = None
            if products_df.empty:
                print(f"  ⚠ Products sheet is empty, cannot resolve product ID")
                continue

            # Try to find by name (case-insensitive)
            prod_match = products_df[
                products_df['name'].astype(str).str.lower() == str(product_name).lower()
            ]

            if prod_match.empty:
                print(f"  ✗ Product not found in Products sheet")
                print(f"  Available products: {list(products_df['name'])}")
                continue

            product_db_id = str(prod_match.iloc[0]['id'])
            print(f"  ✓ Found product ID: {product_db_id}")

            # Find and update stock
            stock_idx = None
            if not stock_df.empty:
                matching_stocks = stock_df[
                    stock_df['product_id'].astype(str) == product_db_id
                ]
                if not matching_stocks.empty:
                    stock_idx = matching_stocks.index[0]

            if stock_idx is not None:
                # Update existing stock
                current_sold = int(stock_df.loc[stock_idx, 'Sold'] or 0)
                current_avail = int(stock_df.loc[stock_idx, 'Available'] or 0)

                stock_df.loc[stock_idx, 'Sold'] = current_sold + qty
                stock_df.loc[stock_idx, 'Available'] = current_avail - qty

                print(f"  ✓ Updated stock: Sold {current_sold}→{current_sold+qty}, Avail {current_avail}→{current_avail-qty}")
                changes_made = True
            else:
                # Create new stock entry
                print(f"  → Creating new stock entry")
                new_row = {
                    'product_id': product_db_id,
                    'product_name': product_name,
                    'Purchased': 0,
                    'Sold': qty,
                    'Available': -qty
                }
                stock_df = pd.concat([
                    stock_df,
                    pd.DataFrame([new_row])
                ], ignore_index=True)
                print(f"  ✓ New entry created with Sold={qty}, Available={-qty}")
                changes_made = True

        # Write back to Excel if changes were made
        if changes_made:
            write_sheet(stock_df, 'Stock')
            print(f"\n✓ Stock sheet updated successfully\n")
            return True
        else:
            print(f"\n⚠ No stock changes were made\n")
            return False

    except Exception as e:
        print(f"\n✗ Error in deduct_stock_for_items: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return False

# ─────────────────────────────────────────────────────────────────
# VENDORS
# ─────────────────────────────────────────────────────────────────
@app.get("/api/vendors")
async def get_vendors():
    df = read_sheet('Vendors')
    if df.empty:
        return []
    return df.to_dict(orient='records')

@app.post("/api/vendors")
async def create_vendor(vendor: VendorModel):
    vendor.id = gen_id()
    df = read_sheet('Vendors')

    if not df.empty:
        if (df['phone'] == vendor.phone).any():
            raise HTTPException(status_code=400, detail="Phone already exists")
        if (df['gstin'] == vendor.gstin).any():
            raise HTTPException(status_code=400, detail="GSTIN already exists")

    new_row = pd.DataFrame([vendor.dict()])
    df = pd.concat([df, new_row], ignore_index=True)
    write_sheet(df, 'Vendors')
    return {"id": vendor.id, "message": "Vendor created"}

@app.put("/api/vendors/{vendor_id}")
async def update_vendor(vendor_id: str, vendor: VendorModel):
    df = read_sheet('Vendors')

    if not (df['id'] == vendor_id).any():
        raise HTTPException(status_code=404, detail="Vendor not found")

    other = df[df['id'] != vendor_id]
    if not other.empty:
        if (other['phone'] == vendor.phone).any():
            raise HTTPException(status_code=400, detail="Phone already used")
        if (other['gstin'] == vendor.gstin).any():
            raise HTTPException(status_code=400, detail="GSTIN already used")

    df.loc[df['id'] == vendor_id] = vendor.dict()
    write_sheet(df, 'Vendors')
    return {"message": "Vendor updated"}

@app.delete("/api/vendors/{vendor_id}")
async def delete_vendor(vendor_id: str):
    df = read_sheet('Vendors')
    df = df[df['id'] != vendor_id]
    write_sheet(df, 'Vendors')
    return {"message": "Vendor deleted"}

@app.get("/api/vendors/{vendor_id}/details")
async def get_vendor_details(vendor_id: str):
    """Get vendor with all their purchase history"""
    vendors_df = read_sheet('Vendors')
    vendor = vendors_df[vendors_df['id'] == vendor_id]

    if vendor.empty:
        raise HTTPException(status_code=404, detail="Vendor not found")

    vendor_row = vendor.iloc[0]
    purchases_df = read_sheet('PurchaseInvoices')

    # Get all purchases for this vendor
    vendor_purchases = []
    if not purchases_df.empty:
        vendor_invoices = purchases_df[purchases_df['vendor_gstin'] == vendor_row['gstin']]
        for _, inv in vendor_invoices.iterrows():
            try:
                products = json.loads(inv['products']) if isinstance(inv['products'], str) else inv['products']
            except:
                products = []

            vendor_purchases.append({
                'id': inv['id'],
                'invoice_no': inv['vendor_invoice_no'],
                'invoice_date': inv['invoice_date'],
                'amount': inv['grand_total'],
                'products': products
            })

    return {
        "id": str(vendor_row['id']),
        "name": str(vendor_row['name']),
        "contact_person": str(vendor_row.get('contact_person', '')),
        "phone": str(vendor_row['phone']),
        "email": str(vendor_row.get('email', '')),
        "gstin": str(vendor_row['gstin']),
        "dl1": str(vendor_row.get('dl1', '')),
        "dl2": str(vendor_row.get('dl2', '')),
        "address": str(vendor_row.get('address', '')),
        "city": str(vendor_row['city']),
        "pincode": str(vendor_row['pincode']),
        "bank_name": str(vendor_row.get('bank_name', '')),
        "bank_acc": str(vendor_row.get('bank_acc', '')),
        "bank_ifsc": str(vendor_row.get('bank_ifsc', '')),
        "purchases": vendor_purchases,
        "total_purchases": int(len(vendor_purchases)),
        "total_amount": float(sum([p['amount'] for p in vendor_purchases]))
    }

# ─────────────────────────────────────────────────────────────────
# PRODUCTS
# ─────────────────────────────────────────────────────────────────
@app.get("/api/products")
async def get_products():
    df = read_sheet('Products')
    if df.empty:
        return []
    return df.to_dict(orient='records')

@app.post("/api/products")
async def create_product(product: ProductModel):
    product.id = gen_id()
    df = read_sheet('Products')

    # Validation: cost_price <= mrp
    if product.cost_price > product.mrp:
        raise HTTPException(status_code=400, detail=f"Cost price (₹{product.cost_price}) cannot exceed MRP (₹{product.mrp})")

    if not df.empty and (df['name'] == product.name).any():
        raise HTTPException(status_code=400, detail="Product name already exists")

    new_row = pd.DataFrame([product.dict()])
    df = pd.concat([df, new_row], ignore_index=True)
    write_sheet(df, 'Products')
    return {"id": product.id, "message": "Product created"}

@app.put("/api/products/{product_id}")
async def update_product(product_id: str, product: ProductModel):
    df = read_sheet('Products')

    if not (df['id'] == product_id).any():
        raise HTTPException(status_code=404, detail="Product not found")

    if product.cost_price > product.mrp:
        raise HTTPException(status_code=400, detail=f"Cost price cannot exceed MRP")

    df.loc[df['id'] == product_id] = product.dict()
    write_sheet(df, 'Products')
    return {"message": "Product updated"}

@app.delete("/api/products/{product_id}")
async def delete_product(product_id: str):
    df = read_sheet('Products')
    df = df[df['id'] != product_id]
    write_sheet(df, 'Products')
    return {"message": "Product deleted"}

# ─────────────────────────────────────────────────────────────────
# PURCHASES
# ─────────────────────────────────────────────────────────────────
@app.get("/api/purchases")
async def get_purchases():
    df = read_sheet('Purchases')
    if df.empty:
        return []

    products_df = read_sheet('Products')
    vendors_df = read_sheet('Vendors')

    if not products_df.empty:
        df = df.merge(products_df[['id', 'name', 'cost_price', 'mrp', 'gst_rate']],
                     left_on='product_id', right_on='id', how='left', suffixes=('', '_prod'))

    if not vendors_df.empty:
        df = df.merge(vendors_df[['id', 'name']],
                     left_on='vendor_id', right_on='id', how='left', suffixes=('', '_vendor'))

    return df.to_dict(orient='records')

@app.post("/api/purchases")
async def create_purchase(purchase: PurchaseModel):
    purchase.id = gen_id()
    purchase.vendor_id = str(purchase.vendor_id)
    purchase.product_id = str(purchase.product_id)
    df = read_sheet('Purchases')

    # Check for duplicate invoice number only if dataframe has data
    if len(df) > 0:
        if purchase.invoice_no in df['invoice_no'].values:
            raise HTTPException(status_code=400, detail="Invoice number already exists")

    new_row = pd.DataFrame([purchase.dict()])
    df = pd.concat([df, new_row], ignore_index=True)
    write_sheet(df, 'Purchases')
    return {"id": purchase.id, "message": "Purchase recorded"}

@app.delete("/api/purchases/{purchase_id}")
async def delete_purchase(purchase_id: str):
    df = read_sheet('Purchases')
    df = df[df['id'] != purchase_id]
    write_sheet(df, 'Purchases')
    return {"message": "Purchase deleted"}

# ─────────────────────────────────────────────────────────────────
# SALES - PRICE VALIDATION: cost_price <= selling_price <= mrp
# ─────────────────────────────────────────────────────────────────
@app.get("/api/sales")
async def get_sales():
    df = read_sheet('Sales')
    if df.empty:
        return []

    # Ensure status column exists and has default value
    if 'status' not in df.columns:
        df['status'] = 'draft'
    else:
        df['status'] = df['status'].fillna('draft')

    products_df = read_sheet('Products')
    if not products_df.empty:
        df = df.merge(products_df[['id', 'name', 'cost_price', 'mrp', 'gst_rate']],
                     left_on='product_id', right_on='id', how='left', suffixes=('', '_product'))

    # Rename id_product back to product_id and keep the sales id as id
    if 'id_product' in df.columns:
        df = df.drop(columns=['id_product'])

    return df.to_dict(orient='records')

# ─────────────────────────────────────────────────────────────────
# DRAFTS (Complete Invoice Data)
# ─────────────────────────────────────────────────────────────────
@app.post("/api/drafts")
async def save_draft(draft: DraftInvoiceModel):
    # Always generate new unique ID for new drafts
    draft.id = gen_id()
    draft.created_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    draft.updated_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    df = read_sheet('Drafts')
    new_row = pd.DataFrame([{
        'id': draft.id,
        'data': json.dumps(draft.data),
        'created_date': draft.created_date,
        'updated_date': draft.updated_date
    }])
    df = pd.concat([df, new_row], ignore_index=True)
    write_sheet(df, 'Drafts')

    return {"id": str(draft.id), "message": "Draft saved"}  # Convert to string to prevent JavaScript precision loss

@app.put("/api/drafts/{draft_id}")
async def update_draft(draft_id: str, draft: DraftInvoiceModel):
    df = read_sheet('Drafts')
    df['id'] = df['id'].astype(str)
    draft_row = df[df['id'] == str(draft_id)]

    if draft_row.empty:
        raise HTTPException(status_code=404, detail="Draft not found")

    draft.updated_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    df.loc[df['id'] == str(draft_id), 'data'] = json.dumps(draft.data)
    df.loc[df['id'] == str(draft_id), 'updated_date'] = draft.updated_date
    write_sheet(df, 'Drafts')

    return {"id": str(draft_id), "message": "Draft updated"}  # Convert to string

@app.delete("/api/drafts/{draft_id}")
async def delete_draft(draft_id: str):
    df = read_sheet('Drafts')
    df['id'] = df['id'].astype(str)
    draft = df[df['id'] == str(draft_id)]

    if draft.empty:
        raise HTTPException(status_code=404, detail="Draft not found")

    df = df[df['id'] != str(draft_id)]
    write_sheet(df, 'Drafts')

    return {"message": "Draft deleted"}

@app.get("/api/drafts")
async def get_drafts():
    df = read_sheet('Drafts')
    if df.empty:
        return []

    result = []
    for idx, row in df.iterrows():
        try:
            data = json.loads(row['data']) if isinstance(row['data'], str) else row['data']
            result.append({
                'id': str(row['id']),  # CRITICAL: Convert to string to prevent JavaScript precision loss
                'data': data,
                'created_date': row['created_date'],
                'updated_date': row['updated_date']
            })
        except:
            continue

    return result

@app.post("/api/drafts/{draft_id}/to-proforma")
async def draft_to_proforma(draft_id: str):
    """Convert draft to proforma (finalize proforma form)"""
    drafts_df = read_sheet('Drafts')
    drafts_df['id'] = drafts_df['id'].astype(str)
    draft = drafts_df[drafts_df['id'] == str(draft_id)]

    if draft.empty:
        raise HTTPException(status_code=404, detail="Draft not found")

    draft_row = draft.iloc[0]
    draft_data = draft_row['data']
    created_date = draft_row['created_date']

    # Ensure data is JSON string, not dict
    if isinstance(draft_data, str):
        data_json = draft_data
    else:
        data_json = json.dumps(draft_data)

    # Move to Proformas sheet
    proformas_df = read_sheet('Proformas')
    new_proforma = pd.DataFrame([{
        'id': draft_id,
        'data': data_json,
        'created_date': created_date,
        'updated_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }])
    proformas_df = pd.concat([proformas_df, new_proforma], ignore_index=True)
    write_sheet(proformas_df, 'Proformas')

    # Delete from Drafts
    drafts_df = drafts_df[drafts_df['id'] != str(draft_id)]
    write_sheet(drafts_df, 'Drafts')

    return {"id": str(draft_id), "message": "Draft converted to proforma"}  # Convert to string

@app.get("/api/proformas")
async def get_proformas():
    """Get all proformas"""
    df = read_sheet('Proformas')
    if df.empty:
        return []

    result = []
    for idx, row in df.iterrows():
        try:
            data = json.loads(row['data']) if isinstance(row['data'], str) else row['data']
            result.append({
                'id': str(row['id']),  # Convert to string to prevent JavaScript precision loss
                'data': data,
                'created_date': row['created_date'],
                'updated_date': row['updated_date']
            })
        except:
            continue

    return result

@app.put("/api/proformas/{proforma_id}")
async def update_proforma(proforma_id: str):
    """Update proforma (currently not used, but available for future use)"""
    df = read_sheet('Proformas')
    df['id'] = df['id'].astype(str)
    proforma = df[df['id'] == str(proforma_id)]

    if proforma.empty:
        raise HTTPException(status_code=404, detail="Proforma not found")

    raise HTTPException(status_code=501, detail="Proforma updates not supported via API")

@app.delete("/api/proformas/{proforma_id}")
async def delete_proforma(proforma_id: str):
    """Delete a proforma"""
    df = read_sheet('Proformas')
    df['id'] = df['id'].astype(str)
    proforma = df[df['id'] == str(proforma_id)]

    if proforma.empty:
        raise HTTPException(status_code=404, detail="Proforma not found")

    df = df[df['id'] != str(proforma_id)]
    write_sheet(df, 'Proformas')

    return {"message": "Proforma deleted"}

@app.post("/api/proformas/{proforma_id}/finalize")
async def finalize_proforma(proforma_id: str, finalize_req: FinalizeModel = None):
    """Finalize proforma to create final invoice"""
    try:
        proformas_df = read_sheet('Proformas')
        if proformas_df.empty:
            raise HTTPException(status_code=404, detail="Proforma not found")

        if 'id' in proformas_df.columns:
            proformas_df['id'] = proformas_df['id'].astype(str)

        proforma = proformas_df[proformas_df['id'] == str(proforma_id)]

        if proforma.empty:
            raise HTTPException(status_code=404, detail="Proforma not found")

        proforma_row = proforma.iloc[0]
        proforma_data = json.loads(proforma_row['data']) if isinstance(proforma_row['data'], str) else proforma_row['data']

        # If frontend provides updated data with the new invoice number, use it
        if finalize_req and finalize_req.data:
            proforma_data = finalize_req.data

        # Stock validation
        purchases_df = read_sheet('Purchases')
        invoices_df = read_sheet('Invoices')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading proforma: {str(e)}")

    # Stock validation - simplified (skip if sheets don't have proper structure)
    # TODO: Implement proper stock validation when Invoices sheet has product_id column
    if 'items' in proforma_data and proforma_data['items']:
        try:
            for item in proforma_data['items']:
                product_id = str(item.get('productId', '') or item.get('product_id', ''))
                qty = int(item.get('qty', 0))
                # Stock validation temporarily skipped
        except:
            pass

    # Create final invoice
    try:
        invoice_id = gen_id()
        finalized_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Read invoices sheet
        invoices_df = read_sheet('Invoices')

        # Expected columns for Invoices sheet
        expected_columns = ['id', 'proforma_id', 'data', 'finalized_date']

        # If sheet is empty, create proper structure with expected columns
        if invoices_df.empty:
            invoices_df = pd.DataFrame(columns=expected_columns)

        # Create new invoice row
        new_invoice = pd.DataFrame([{
            'id': str(invoice_id),
            'proforma_id': str(proforma_id),
            'data': json.dumps(proforma_data),
            'finalized_date': finalized_date
        }])

        # Ensure column order matches expected
        if not invoices_df.empty:
            # Reorder columns to match expected order
            invoices_df = invoices_df[expected_columns]
            new_invoice = new_invoice[expected_columns]
            invoices_df = pd.concat([invoices_df, new_invoice], ignore_index=True)
        else:
            # Use new_invoice as-is if invoices_df is empty
            invoices_df = new_invoice[expected_columns]

        # Write to sheet
        write_sheet(invoices_df, 'Invoices')

        # Remove from Proformas sheet
        proformas_df = proformas_df[proformas_df['id'] != str(proforma_id)]
        write_sheet(proformas_df, 'Proformas')

        # Deduct stock for items in the invoice
        items_list = proforma_data.get('items', []) if 'items' in proforma_data else []
        if items_list:
            deduct_stock_for_items(items_list)
        else:
            print(f"⚠ No items in invoice data, skipping stock deduction")

        return {"invoice_id": str(invoice_id), "invNo": proforma_data.get('invNo', ''), "message": "Invoice finalized, stock deducted"}
    except Exception as e:
        import traceback
        error_msg = f"Error creating invoice: {str(e)}\n{traceback.format_exc()}"
        raise HTTPException(status_code=500, detail=error_msg)

@app.get("/api/next-invoice-number/{invoice_type}")
async def get_next_invoice_number(invoice_type: str):
    """Get next sequential invoice number and increment counter"""
    df = read_sheet('Counters')

    # Map invoice types to counter names
    counter_map = {
        'retail': 'retail_invoice',
        'patient': 'patient_invoice',
        'purchase': 'purchase_invoice'
    }

    counter_name = counter_map.get(invoice_type.lower())
    if not counter_name:
        raise HTTPException(status_code=400, detail=f"Invalid invoice type: {invoice_type}")

    # Get current counter value
    counter_row = df[df['counter_name'] == counter_name]
    if counter_row.empty:
        # Initialize if doesn't exist
        current = 0
    else:
        current = int(counter_row.iloc[0]['current_value'])

    # Increment counter
    next_num = current + 1

    # Update in sheet
    if counter_row.empty:
        new_row = pd.DataFrame([{'counter_name': counter_name, 'current_value': next_num}])
        df = pd.concat([df, new_row], ignore_index=True)
    else:
        df.loc[df['counter_name'] == counter_name, 'current_value'] = next_num

    write_sheet(df, 'Counters')

    # Generate invoice number based on type
    year = datetime.now().year
    if invoice_type.lower() == 'retail':
        invoice_no = f'GSCR-{year}-{str(next_num).zfill(4)}'
    elif invoice_type.lower() == 'patient':
        invoice_no = f'GCPC-{year}-{str(next_num).zfill(4)}'
    elif invoice_type.lower() == 'purchase':
        invoice_no = f'PO-{year}-{str(next_num).zfill(5)}'

    return {"invoice_number": invoice_no, "counter": next_num}

@app.get("/api/invoices")
async def get_invoices():
    df = read_sheet('Invoices')
    if df.empty:
        return []

    result = []
    for idx, row in df.iterrows():
        try:
            data = json.loads(row['data']) if isinstance(row['data'], str) else row['data']
            result.append({
                'id': str(row['id']),  # Convert to string to prevent JavaScript precision loss
                'proforma_id': row.get('proforma_id', ''),
                'data': data,
                'finalized_date': row.get('finalized_date', '')
            })
        except:
            continue

    return result

# ─────────────────────────────────────────────────────────────────
# PURCHASE INVOICES (Vendor Purchases)

@app.post("/api/purchase-invoices")
async def create_purchase_invoice(purchase: PurchaseInvoiceModel):
    # Check for duplicate vendor invoice number
    df = read_sheet('PurchaseInvoices')
    duplicate = df[(df['vendor_gstin'] == purchase.vendor_gstin) & (df['vendor_invoice_no'] == purchase.vendor_invoice_no)]
    if not duplicate.empty:
        raise HTTPException(status_code=400, detail=f"Invoice number {purchase.vendor_invoice_no} already exists for this vendor")

    # Auto-create vendor if not exists
    vendors_df = read_sheet('Vendors')
    vendor_exists = (vendors_df['gstin'] == purchase.vendor_gstin).any() if not vendors_df.empty else False

    if not vendor_exists:
        vendor_id = gen_id()
        new_vendor = pd.DataFrame([{
            'id': vendor_id,
            'name': purchase.vendor_name,
            'contact_person': purchase.contact_person if hasattr(purchase, 'contact_person') else '',
            'phone': purchase.phone,
            'gstin': purchase.vendor_gstin,
            'dl1': purchase.dl_no_1 if hasattr(purchase, 'dl_no_1') else '',
            'dl2': purchase.dl_no_2 if hasattr(purchase, 'dl_no_2') else '',
            'address': purchase.address if hasattr(purchase, 'address') else '',
            'city': purchase.city,
            'pincode': purchase.pincode,
            'bank_name': purchase.vendor_bank_name if hasattr(purchase, 'vendor_bank_name') else '',
            'bank_acc': purchase.vendor_account_no if hasattr(purchase, 'vendor_account_no') else '',
            'bank_ifsc': purchase.vendor_ifsc if hasattr(purchase, 'vendor_ifsc') else ''
        }])
        vendors_df = pd.concat([vendors_df, new_vendor], ignore_index=True)
        write_sheet(vendors_df, 'Vendors')

    purchase.id = gen_id()
    purchase.created_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    purchase.updated_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    df = read_sheet('PurchaseInvoices')
    new_row = pd.DataFrame([{
        'id': purchase.id,
        'vendor_name': purchase.vendor_name,
        'vendor_gstin': purchase.vendor_gstin,
        'phone': purchase.phone,
        'city': purchase.city,
        'pincode': purchase.pincode,
        'vendor_invoice_no': purchase.vendor_invoice_no,
        'invoice_date': purchase.invoice_date,
        'payment_mode': purchase.payment_mode,
        'amount_paid': purchase.amount_paid,
        'paid_by': purchase.paid_by,
        'products': json.dumps(purchase.products),
        'subtotal': purchase.subtotal,
        'total_gst': purchase.total_gst,
        'discount_amount': purchase.discount_amount,
        'grand_total': purchase.grand_total,
        'created_date': purchase.created_date,
        'updated_date': purchase.updated_date
    }])
    df = pd.concat([df, new_row], ignore_index=True)
    write_sheet(df, 'PurchaseInvoices')

    # Auto-create products if they don't exist
    try:
        products_df = read_sheet('Products')
    except:
        products_df = pd.DataFrame()

    for product in purchase.products:
        product_id = str(product.get('product_id', ''))
        if product_id and not products_df.empty:
            if not (products_df['id'].astype(str) == product_id).any():
                # Create new product
                new_product = pd.DataFrame([{
                    'id': product_id,
                    'name': product.get('name', ''),
                    'hsn_code': product.get('hsn', ''),
                    'pack': product.get('pack', ''),
                    'company': product.get('company', ''),
                    'scheme': product.get('scheme', ''),
                    'cost_price': float(product.get('buy_rate', 0)),
                    'mrp': float(product.get('mrp', 0)),
                    'gst_rate': float(product.get('gst', 0))
                }])
                products_df = pd.concat([products_df, new_product], ignore_index=True)
        elif product_id and products_df.empty:
            # Create first product
            products_df = pd.DataFrame([{
                'id': product_id,
                'name': product.get('name', ''),
                'hsn_code': product.get('hsn', ''),
                'pack': product.get('pack', ''),
                'company': product.get('company', ''),
                'scheme': product.get('scheme', ''),
                'cost_price': float(product.get('buy_rate', 0)),
                'mrp': float(product.get('mrp', 0)),
                'gst_rate': float(product.get('gst', 0))
            }])

    if not products_df.empty:
        write_sheet(products_df, 'Products')

    # Auto-add products to stock
    try:
        stock_df = read_sheet('Stock')
    except:
        stock_df = pd.DataFrame()

    for product in purchase.products:
        product_id = str(product.get('product_id', ''))
        product_name = product.get('name', '')
        qty = int(product.get('qty', 0))

        # Add to Purchases sheet
        purchases_df = read_sheet('Purchases')
        new_purchase = pd.DataFrame([{
            'id': gen_id(),
            'vendor_id': 'vendor_' + purchase.vendor_gstin[:8],
            'product_id': product_id,
            'quantity': qty,
            'batch_no': product.get('batch', ''),
            'expiry_date': product.get('expiry', ''),
            'invoice_no': purchase.vendor_invoice_no,
            'invoice_date': purchase.invoice_date,
            'amount_paid': float(product.get('qty', 0)) * float(product.get('buy_rate', 0)),
            'paid_by': purchase.paid_by
        }])
        purchases_df = pd.concat([purchases_df, new_purchase], ignore_index=True)
        write_sheet(purchases_df, 'Purchases')

        # Update Stock sheet - increase Purchased and Available
        if product_id and qty > 0:
            product_rows = stock_df[stock_df['product_id'] == product_id]
            if not product_rows.empty:
                # Update existing product in stock
                idx = product_rows.index[0]
                current_purchased = int(stock_df.loc[idx, 'Purchased'] or 0)
                current_available = int(stock_df.loc[idx, 'Available'] or 0)
                stock_df.loc[idx, 'Purchased'] = current_purchased + qty
                stock_df.loc[idx, 'Available'] = current_available + qty
            else:
                # Add new product to stock
                new_stock_row = pd.DataFrame([{
                    'product_id': product_id,
                    'product_name': product_name,
                    'Purchased': qty,
                    'Sold': 0,
                    'Available': qty
                }])
                stock_df = pd.concat([stock_df, new_stock_row], ignore_index=True)

    # Write updated stock sheet
    if not stock_df.empty:
        write_sheet(stock_df, 'Stock')

    # Add to Expiry sheet
    try:
        expiry_df = read_sheet('Expiry')
    except:
        expiry_df = pd.DataFrame()

    for product in purchase.products:
        product_name = product.get('name', '')
        batch = product.get('batch', '')
        expiry = product.get('expiry', '')
        qty = int(product.get('qty', 0))
        product_id = str(product.get('product_id', ''))

        if expiry and batch and qty > 0:
            new_expiry = pd.DataFrame([{
                'product_id': product_id,
                'product_name': product_name,
                'batch': batch,
                'expiry_date': expiry,
                'quantity': qty,
                'purchase_id': purchase.id
            }])
            expiry_df = pd.concat([expiry_df, new_expiry], ignore_index=True)

    if not expiry_df.empty:
        write_sheet(expiry_df, 'Expiry')

    return {"id": purchase.id, "message": "Purchase invoice created, vendor recorded, stock and expiry updated"}

@app.get("/api/purchase-invoices")
async def get_purchase_invoices():
    df = read_sheet('PurchaseInvoices')
    if df.empty:
        return []

    result = []
    for idx, row in df.iterrows():
        try:
            products = json.loads(row['products']) if isinstance(row['products'], str) else row['products']
            result.append({
                'id': str(row['id']),  # Convert to string to prevent JavaScript precision loss
                'vendor_name': row.get('vendor_name', ''),
                'vendor_gstin': row.get('vendor_gstin', ''),
                'phone': row.get('phone', ''),
                'city': row.get('city', ''),
                'pincode': row.get('pincode', ''),
                'vendor_invoice_no': row.get('vendor_invoice_no', ''),
                'invoice_date': row.get('invoice_date', ''),
                'payment_mode': row.get('payment_mode', ''),
                'amount_paid': row.get('amount_paid', 0),
                'paid_by': row.get('paid_by', ''),
                'products': products,
                'subtotal': row.get('subtotal', 0),
                'total_gst': row.get('total_gst', 0),
                'discount_amount': row.get('discount_amount', 0),
                'grand_total': row.get('grand_total', 0),
                'created_date': row.get('created_date', '')
            })
        except:
            continue

    return result

@app.put("/api/purchase-invoices/{invoice_id}")
async def update_purchase_invoice(invoice_id: str, purchase: PurchaseInvoiceModel):
    df = read_sheet('PurchaseInvoices')
    df['id'] = df['id'].astype(str)
    invoice = df[df['id'] == str(invoice_id)]

    if invoice.empty:
        raise HTTPException(status_code=404, detail="Purchase invoice not found")

    purchase.updated_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    df.loc[df['id'] == str(invoice_id), 'vendor_name'] = purchase.vendor_name
    df.loc[df['id'] == str(invoice_id), 'products'] = json.dumps(purchase.products)
    df.loc[df['id'] == str(invoice_id), 'subtotal'] = purchase.subtotal
    df.loc[df['id'] == str(invoice_id), 'total_gst'] = purchase.total_gst
    df.loc[df['id'] == str(invoice_id), 'discount_amount'] = purchase.discount_amount
    df.loc[df['id'] == str(invoice_id), 'grand_total'] = purchase.grand_total
    df.loc[df['id'] == str(invoice_id), 'updated_date'] = purchase.updated_date
    write_sheet(df, 'PurchaseInvoices')

    return {"id": invoice_id, "message": "Purchase invoice updated"}

@app.delete("/api/purchase-invoices/{invoice_id}")
async def delete_purchase_invoice(invoice_id: str):
    df = read_sheet('PurchaseInvoices')
    df['id'] = df['id'].astype(str)
    invoice = df[df['id'] == str(invoice_id)]

    if invoice.empty:
        raise HTTPException(status_code=404, detail="Purchase invoice not found")

    df = df[df['id'] != str(invoice_id)]
    write_sheet(df, 'PurchaseInvoices')

    return {"message": "Purchase invoice deleted"}

@app.post("/api/sales")
async def create_sale(sale: SalesModel):
    sale.id = gen_id()
    sale.status = "draft"

    products_df = read_sheet('Products')
    products_df['id'] = products_df['id'].astype(str)
    product = products_df[products_df['id'] == str(sale.product_id)]

    if product.empty:
        raise HTTPException(status_code=404, detail="Product not found")

    cost_price = float(product.iloc[0]['cost_price'])
    mrp = float(product.iloc[0]['mrp'])

    # PRICE VALIDATION
    if sale.selling_price < cost_price:
        raise HTTPException(status_code=400, detail=f"❌ Cannot sell below COST PRICE ₹{cost_price}")

    if sale.selling_price > mrp:
        raise HTTPException(status_code=400, detail=f"❌ Cannot sell above MRP ₹{mrp}")

    # Draft saved without stock check - stock only deducts on finalization
    df = read_sheet('Sales')
    new_row = pd.DataFrame([sale.dict()])
    df = pd.concat([df, new_row], ignore_index=True)
    write_sheet(df, 'Sales')

    return {"id": sale.id, "message": "Draft created"}

@app.put("/api/sales/{sale_id}")
async def update_sale(sale_id: str, sale: SalesModel):
    df = read_sheet('Sales')
    df['id'] = df['id'].astype(str)
    draft = df[df['id'] == str(sale_id)]

    if draft.empty:
        raise HTTPException(status_code=404, detail="Draft not found")

    if draft.iloc[0]['status'] == 'final':
        raise HTTPException(status_code=400, detail="Cannot edit finalized invoices")

    products_df = read_sheet('Products')
    products_df['id'] = products_df['id'].astype(str)
    product = products_df[products_df['id'] == str(sale.product_id)]

    if product.empty:
        raise HTTPException(status_code=404, detail="Product not found")

    cost_price = float(product.iloc[0]['cost_price'])
    mrp = float(product.iloc[0]['mrp'])

    if sale.selling_price < cost_price:
        raise HTTPException(status_code=400, detail=f"Cannot sell below cost price ₹{cost_price}")

    if sale.selling_price > mrp:
        raise HTTPException(status_code=400, detail=f"Cannot sell above MRP ₹{mrp}")

    sale.status = "draft"
    df.loc[df['id'] == sale_id] = sale.dict()
    write_sheet(df, 'Sales')
    return {"message": "Draft updated"}

@app.delete("/api/sales/{sale_id}")
async def delete_sale(sale_id: str):
    df = read_sheet('Sales')
    df['id'] = df['id'].astype(str)
    draft = df[df['id'] == str(sale_id)]

    if draft.empty:
        raise HTTPException(status_code=404, detail="Draft not found")

    if draft.iloc[0]['status'] == 'final':
        raise HTTPException(status_code=400, detail="Cannot delete finalized invoices")

    df = df[df['id'] != sale_id]
    write_sheet(df, 'Sales')
    return {"message": "Draft deleted"}

@app.post("/api/sales/{sale_id}/finalize")
async def finalize_sale(sale_id: str):
    sales_df = read_sheet('Sales')
    sales_df['id'] = sales_df['id'].astype(str)
    draft = sales_df[sales_df['id'] == str(sale_id)]

    if draft.empty:
        raise HTTPException(status_code=404, detail="Draft not found")

    draft_row = draft.iloc[0]

    if draft_row['status'] == 'final':
        raise HTTPException(status_code=400, detail="Invoice already finalized")

    purchases_df = read_sheet('Purchases')
    invoices_df = read_sheet('Invoices')

    purchases_df['product_id'] = purchases_df['product_id'].astype(str)
    invoices_df['product_id'] = invoices_df['product_id'].astype(str)

    product_id = str(draft_row['product_id'])
    quantity = int(draft_row['quantity'])

    total_purchased = purchases_df[purchases_df['product_id'] == product_id]['quantity'].sum() if not purchases_df.empty else 0
    total_invoiced = invoices_df[invoices_df['product_id'] == product_id]['quantity'].sum() if not invoices_df.empty else 0
    available = total_purchased - total_invoiced

    if quantity > available:
        raise HTTPException(status_code=400, detail=f"❌ Stock insufficient. Available: {available}")

    invoice_id = gen_id()
    finalized_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    invoice_data = {
        'id': invoice_id,
        'draft_id': sale_id,
        'customer_name': draft_row['customer_name'],
        'customer_phone': draft_row['customer_phone'],
        'product_id': draft_row['product_id'],
        'quantity': draft_row['quantity'],
        'selling_price': draft_row['selling_price'],
        'sale_date': draft_row['sale_date'],
        'finalized_date': finalized_date
    }

    new_invoice = pd.DataFrame([invoice_data])
    invoices_df = pd.concat([invoices_df, new_invoice], ignore_index=True)
    write_sheet(invoices_df, 'Invoices')

    sales_df.loc[sales_df['id'] == sale_id, 'status'] = 'final'
    write_sheet(sales_df, 'Sales')

    return {"invoice_id": invoice_id, "message": "Invoice finalized, stock deducted"}

# ─────────────────────────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────────────────────────
@app.post("/api/reset-stock")
async def reset_stock():
    """Reset stock to original purchased amounts (zero out sold/available adjustments)"""
    try:
        stock_df = read_sheet('Stock')

        if stock_df.empty:
            return {"message": "No stock data to reset"}

        # Reset Sold to 0 for all products
        stock_df['Sold'] = 0

        # Set Available equal to Purchased
        if 'Purchased' in stock_df.columns:
            stock_df['Available'] = stock_df['Purchased']

        write_sheet(stock_df, 'Stock')
        return {"message": "Stock reset successfully. All sales deductions cleared.", "stock_data": stock_df.to_dict(orient='records')}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error resetting stock: {str(e)}")

@app.delete("/api/invoices/{invoice_id}")
async def delete_invoice(invoice_id: str):
    """Delete a finalized invoice and restore stock"""
    try:
        invoices_df = read_sheet('Invoices')

        if invoices_df.empty:
            raise HTTPException(status_code=404, detail="Invoice not found")

        # Find the invoice to delete
        invoice_rows = invoices_df[invoices_df['id'].astype(str) == str(invoice_id)]
        if invoice_rows.empty:
            raise HTTPException(status_code=404, detail="Invoice not found")

        invoice_data = json.loads(invoice_rows.iloc[0]['data'])

        # Restore stock (reverse the deduction)
        items_list = invoice_data.get('items', [])
        if items_list:
            try:
                stock_df = read_sheet('Stock')
                products_df = read_sheet('Products')

                for item in items_list:
                    product_name = item.get('name', '')
                    qty = int(item.get('qty', 0))

                    if not product_name or qty <= 0:
                        continue

                    # Find product in Products sheet
                    prod_match = products_df[
                        products_df['name'].astype(str).str.lower() == str(product_name).lower()
                    ]

                    if prod_match.empty:
                        continue

                    product_db_id = str(prod_match.iloc[0]['id'])

                    # Find and restore stock
                    if not stock_df.empty:
                        matching_stocks = stock_df[
                            stock_df['product_id'].astype(str) == product_db_id
                        ]
                        if not matching_stocks.empty:
                            stock_idx = matching_stocks.index[0]
                            current_sold = int(stock_df.loc[stock_idx, 'Sold'] or 0)
                            current_avail = int(stock_df.loc[stock_idx, 'Available'] or 0)

                            stock_df.loc[stock_idx, 'Sold'] = current_sold - qty
                            stock_df.loc[stock_idx, 'Available'] = current_avail + qty
                            print(f"Restored stock for {product_name}: Sold {current_sold}→{current_sold-qty}, Avail {current_avail}→{current_avail+qty}")

                # Write updated stock sheet
                write_sheet(stock_df, 'Stock')
                print(f"Stock restored for invoice {invoice_id}")
            except Exception as e:
                print(f"Warning: Could not restore stock: {str(e)}")

        # Remove from Invoices sheet
        invoices_df = invoices_df[invoices_df['id'].astype(str) != str(invoice_id)]
        write_sheet(invoices_df, 'Invoices')

        return {"message": f"Invoice {invoice_id} deleted successfully. Stock restored."}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error deleting invoice: {str(e)}")

@app.get("/api/dashboard")
async def get_dashboard():
    try:
        vendors_df = read_sheet('Vendors')
        products_df = read_sheet('Products')
        purchases_df = read_sheet('Purchases')
        invoices_df = read_sheet('Invoices')
        stock_df = read_sheet('Stock')

        total_vendors = len(vendors_df) if not vendors_df.empty else 0
        total_products = len(products_df) if not products_df.empty else 0

        # Calculate total purchases
        total_purchases = 0.0
        if not purchases_df.empty and 'amount_paid' in purchases_df.columns:
            total_purchases = float(purchases_df['amount_paid'].sum())

        # Calculate total sales from finalized invoices (stored as JSON in 'data' column)
        total_sales = 0.0
        if not invoices_df.empty and 'data' in invoices_df.columns:
            for _, row in invoices_df.iterrows():
                try:
                    invoice_data = json.loads(row['data']) if isinstance(row['data'], str) else row['data']
                    # Get the net amount (total after tax) or subtotal
                    invoice_amount = float(invoice_data.get('net', invoice_data.get('sub', 0)))
                    total_sales += invoice_amount
                except:
                    pass

        # Calculate total stock from Stock sheet
        total_stock = 0
        if not stock_df.empty and 'Available' in stock_df.columns:
            total_stock = int(stock_df['Available'].sum())
        else:
            # Fallback: calculate from Purchases - Sales
            if not purchases_df.empty and 'quantity' in purchases_df.columns:
                total_stock = int(purchases_df['quantity'].sum())

        expiring_soon = 0

        return {
            "total_vendors": int(total_vendors),
            "total_products": int(total_products),
            "total_purchases": round(total_purchases, 2),
            "total_sales": round(total_sales, 2),
            "total_stock": int(total_stock),
            "expiring_soon": int(expiring_soon)
        }
    except Exception as e:
        import traceback
        return {"error": str(e), "trace": traceback.format_exc()}

@app.get("/api/stock")
async def get_stock():
    try:
        products_df = read_sheet('Products')
        stock_df = read_sheet('Stock')
        purchases_df = read_sheet('Purchases')

        stock_data = []
        if not products_df.empty:
            for _, prod in products_df.iterrows():
                prod_id = str(prod['id'])

                # Try to get stock from Stock sheet first
                purchased = 0
                sold = 0
                available = 0

                if not stock_df.empty:
                    stock_rows = stock_df[stock_df['product_id'].astype(str) == prod_id]
                    if not stock_rows.empty:
                        purchased = int(stock_rows['Purchased'].iloc[0] or 0)
                        sold = int(stock_rows['Sold'].iloc[0] or 0)
                        available = int(stock_rows['Available'].iloc[0] or 0)

                # Get latest batch and expiry from purchases for this product
                batch_no = 'BATCH-001'
                expiry_date = ''
                if not purchases_df.empty and 'product_id' in purchases_df.columns:
                    try:
                        prod_purchases = purchases_df[purchases_df['product_id'].astype(str) == prod_id]
                        if not prod_purchases.empty:
                            batch_no = prod_purchases['batch_no'].iloc[-1] if 'batch_no' in prod_purchases.columns else 'BATCH-001'
                            expiry_date = prod_purchases['expiry_date'].iloc[-1] if 'expiry_date' in prod_purchases.columns else ''
                    except:
                        pass

                stock_data.append({
                    "product_id": prod['id'],
                    "product_name": prod['name'],
                    "hsn_code": prod.get('hsn_code', ''),
                    "pack": prod.get('pack', ''),
                    "company": prod.get('company', ''),
                    "scheme": prod.get('scheme', ''),
                    "cost_price": prod['cost_price'],
                    "mrp": prod['mrp'],
                    "gst_rate": prod['gst_rate'],
                    "batch_no": batch_no,
                    "expiry_date": expiry_date,
                    "purchased": purchased,
                    "sold": sold,
                    "available": available
                })

        return stock_data
    except Exception as e:
        return {"error": str(e)}

@app.post("/api/stock/add")
async def add_stock(stock: StockAddModel):
    """Add stock with vendor details and expiry tracking"""
    try:
        stock_df = read_sheet('Stock')
        expiry_df = read_sheet('Expiry')
        purchases_df = read_sheet('Purchases')

        # Check if stock already exists for this product
        prod_stock = stock_df[stock_df['product_id'].astype(str) == stock.product_id]

        if not prod_stock.empty:
            # Update existing stock
            purchased = int(prod_stock['Purchased'].iloc[0] or 0) + stock.quantity
            available = int(prod_stock['Available'].iloc[0] or 0) + stock.quantity
            stock_df.loc[stock_df['product_id'].astype(str) == stock.product_id, 'Purchased'] = purchased
            stock_df.loc[stock_df['product_id'].astype(str) == stock.product_id, 'Available'] = available
        else:
            # Create new stock entry
            new_stock = {
                'product_id': stock.product_id,
                'product_name': stock.product_name,
                'Purchased': stock.quantity,
                'Sold': 0,
                'Available': stock.quantity
            }
            stock_df = pd.concat([stock_df, pd.DataFrame([new_stock])], ignore_index=True)

        write_sheet(stock_df, 'Stock')

        # Add to Expiry tracking
        purchase_id = gen_id()
        expiry_entry = {
            'product_id': stock.product_id,
            'product_name': stock.product_name,
            'batch': stock.batch_no,
            'expiry_date': stock.expiry_date,
            'quantity': stock.quantity,
            'purchase_id': purchase_id
        }
        expiry_df = pd.concat([expiry_df, pd.DataFrame([expiry_entry])], ignore_index=True)
        write_sheet(expiry_df, 'Expiry')

        # Create purchase record
        purchase_record = {
            'id': purchase_id,
            'vendor_id': stock.vendor_id or '',
            'product_id': stock.product_id,
            'quantity': stock.quantity,
            'batch_no': stock.batch_no,
            'expiry_date': stock.expiry_date,
            'invoice_no': stock.invoice_no or '',
            'invoice_date': stock.invoice_date or datetime.now().strftime("%Y-%m-%d"),
            'amount_paid': stock.amount_paid or 0.0,
            'paid_by': stock.paid_by or ''
        }
        purchases_df = pd.concat([purchases_df, pd.DataFrame([purchase_record])], ignore_index=True)
        write_sheet(purchases_df, 'Purchases')

        return {
            "id": purchase_id,
            "message": f"Stock added: {stock.quantity} units of {stock.product_name}",
            "batch": stock.batch_no,
            "expiry": stock.expiry_date
        }
    except Exception as e:
        import traceback
        raise HTTPException(status_code=400, detail=f"Error adding stock: {str(e)}")

@app.put("/api/stock/{product_id}")
async def update_stock(product_id: str, stock: StockAddModel):
    """Update stock for a product"""
    try:
        stock_df = read_sheet('Stock')

        prod_stock = stock_df[stock_df['product_id'].astype(str) == product_id]
        if prod_stock.empty:
            raise HTTPException(status_code=404, detail="Product stock not found")

        # Update quantities
        stock_df.loc[stock_df['product_id'].astype(str) == product_id, 'Purchased'] = stock.quantity
        available = stock.quantity - int(prod_stock['Sold'].iloc[0] or 0)
        stock_df.loc[stock_df['product_id'].astype(str) == product_id, 'Available'] = available

        write_sheet(stock_df, 'Stock')

        return {"message": f"Stock updated: {stock.quantity} units"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error updating stock: {str(e)}")

@app.post("/api/test/deduct-stock")
async def test_deduct_stock(items: list = None):
    """TEST endpoint to debug stock deduction"""
    if not items:
        items = [
            {"name": "PARACETAMOL 650MG", "qty": 2}
        ]

    print(f"\n{'='*70}")
    print(f"TEST: Deducting stock for items: {items}")
    print(f"{'='*70}")
    result = deduct_stock_for_items(items)

    # Return current stock status
    stock_after = await get_stock()
    return {
        "test_status": "completed",
        "deduction_result": result,
        "stock_after": stock_after
    }

@app.get("/api/states")
async def get_states():
    """Get all states with their SGST/CGST rates"""
    try:
        states_df = read_sheet('States')
        if states_df.empty:
            return []

        states_list = []
        for _, row in states_df.iterrows():
            states_list.append({
                "state_code": str(row['state_code']),
                "state_name": str(row['state_name']),
                "sgst_rate": float(row['sgst_rate']),
                "cgst_rate": float(row['cgst_rate']),
                "igst_rate": float(row['igst_rate'])
            })
        return states_list
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/states/{state_code}")
async def get_state(state_code: str):
    """Get specific state tax rates"""
    try:
        states_df = read_sheet('States')
        state = states_df[states_df['state_code'].astype(str) == state_code.upper()]

        if state.empty:
            raise HTTPException(status_code=404, detail=f"State {state_code} not found")

        row = state.iloc[0]
        return {
            "state_code": str(row['state_code']),
            "state_name": str(row['state_name']),
            "sgst_rate": float(row['sgst_rate']),
            "cgst_rate": float(row['cgst_rate']),
            "igst_rate": float(row['igst_rate'])
        }
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/expiry")
async def get_expiry():
    try:
        purchases_df = read_sheet('Purchases')
        products_df = read_sheet('Products')
        stock_df = read_sheet('Stock')

        if purchases_df.empty:
            return []

        purchases_df['expiry_date'] = pd.to_datetime(purchases_df['expiry_date'])
        if not products_df.empty:
            purchases_df = purchases_df.merge(products_df[['id', 'name']],
                                             left_on='product_id', right_on='id', how='left')

        from datetime import timedelta
        today = datetime.now().date()

        expiry_data = []
        for _, row in purchases_df.iterrows():
            prod_id = str(row['product_id'])
            expiry = row['expiry_date'].date()
            if expiry < today:
                status = "expired"
            elif expiry < today + timedelta(days=30):
                status = "expiring_soon"
            elif expiry < today + timedelta(days=90):
                status = "expiring_3m"
            else:
                status = "good"

            # Get available quantity from Stock sheet instead of raw purchase quantity
            available_qty = row['quantity']  # default to purchased quantity
            if not stock_df.empty:
                stock_rows = stock_df[stock_df['product_id'].astype(str) == prod_id]
                if not stock_rows.empty:
                    available_qty = int(stock_rows['Available'].iloc[0] or 0)

            expiry_data.append({
                "product_id": row['product_id'],
                "product_name": row['name'],
                "batch_no": row['batch_no'],
                "quantity": available_qty,
                "expiry_date": str(row['expiry_date'].date()),
                "status": status
            })

        return expiry_data
    except Exception as e:
        return {"error": str(e)}


@app.get("/")
async def root():
    """Serve the frontend HTML"""
    html_file = Path(__file__).parent.parent / "graftcare.html"
    if html_file.exists():
        return FileResponse(html_file, media_type="text/html")
    return {"message": "Graft Care Pharma API"}

@app.get("/api/health")
async def health():
    return {"status": "running"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
