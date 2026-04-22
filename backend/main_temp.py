from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional, List
import pandas as pd
import json
from pathlib import Path
from datetime import datetime
import os
from openpyxl import load_workbook, Workbook

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Data paths
DATA_DIR = Path(__file__).parent.parent / "data"
DATA_DIR.mkdir(exist_ok=True)

EXCEL_FILE = DATA_DIR / "graftcare.xlsx"

# Models
class VendorModel(BaseModel):
    id: Optional[str] = None
    name: str
    contact_person: str
    phone: str
    gstin: str
    dl1: str
    dl2: str
    address: str
    city: str
    pincode: str
    bank_name: str
    bank_acc: str
    bank_ifsc: str

class ProductModel(BaseModel):
    id: Optional[str] = None
    name: str
    hsn_code: str
    cost_price: float
    mrp: float
    gst_rate: float

class PurchaseModel(BaseModel):
    id: Optional[str] = None
    vendor_id: str
    product_id: str
    quantity: int
    batch_no: str
    expiry_date: str
    invoice_no: str
    invoice_date: str
    amount_paid: float
    paid_by: str

class SalesModel(BaseModel):
    id: Optional[str] = None
    customer_name: str
    customer_phone: str
    product_id: str
    quantity: int
    selling_price: float
    sale_date: str

def gen_id():
    return datetime.now().strftime("%Y%m%d%H%M%S") + str(abs(hash(os.urandom(4))) % 10000)

def init_excel():
    """Create Excel file with sheets"""
    if not EXCEL_FILE.exists():
        wb = Workbook()
        wb.remove(wb.active)

        # Vendors sheet
        ws = wb.create_sheet("Vendors")
        ws.append(['id','name','contact_person','phone','gstin','dl1','dl2','address','city','pincode','bank_name','bank_acc','bank_ifsc'])

        # Products sheet
        ws = wb.create_sheet("Products")
        ws.append(['id','name','hsn_code','cost_price','mrp','gst_rate'])

        # Purchases sheet
        ws = wb.create_sheet("Purchases")
        ws.append(['id','vendor_id','product_id','quantity','batch_no','expiry_date','invoice_no','invoice_date','amount_paid','paid_by'])

        # Sales sheet
        ws = wb.create_sheet("Sales")
        ws.append(['id','customer_name','customer_phone','product_id','quantity','selling_price','sale_date'])

        wb.save(EXCEL_FILE)

init_excel()

def read_sheet(sheet_name):
    """Read Excel sheet"""
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
        df = df.fillna('')
        return df
    except:
        return pd.DataFrame()

def write_sheet(df, sheet_name):
    """Write to Excel sheet"""
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

# ─────────────────────────────────────────────────────────────────
# VENDORS
# ─────────────────────────────────────────────────────────────────
@app.get("/api/vendors")
async def get_vendors():
    df = read_sheet('Vendors')
    if df.empty:
        return []
    return df.to_dict(orient='records')

@app.post("/api/vendors")
async def create_vendor(vendor: VendorModel):
    vendor.id = gen_id()
    df = read_sheet('Vendors')

    if not df.empty:
        if (df['phone'] == vendor.phone).any():
            raise HTTPException(status_code=400, detail="Phone already exists")
        if (df['gstin'] == vendor.gstin).any():
            raise HTTPException(status_code=400, detail="GSTIN already exists")

    new_row = pd.DataFrame([vendor.dict()])
    df = pd.concat([df, new_row], ignore_index=True)
    write_sheet(df, 'Vendors')
    return {"id": vendor.id, "message": "Vendor created"}

@app.put("/api/vendors/{vendor_id}")
async def update_vendor(vendor_id: str, vendor: VendorModel):
    df = read_sheet('Vendors')

    if not (df['id'] == vendor_id).any():
        raise HTTPException(status_code=404, detail="Vendor not found")

    other = df[df['id'] != vendor_id]
    if not other.empty:
        if (other['phone'] == vendor.phone).any():
            raise HTTPException(status_code=400, detail="Phone already used")
        if (other['gstin'] == vendor.gstin).any():
            raise HTTPException(status_code=400, detail="GSTIN already used")

    df.loc[df['id'] == vendor_id] = vendor.dict()
    write_sheet(df, 'Vendors')
    return {"message": "Vendor updated"}

@app.delete("/api/vendors/{vendor_id}")
async def delete_vendor(vendor_id: str):
    df = read_sheet('Vendors')
    df = df[df['id'] != vendor_id]
    write_sheet(df, 'Vendors')
    return {"message": "Vendor deleted"}

# ─────────────────────────────────────────────────────────────────
# PRODUCTS
# ─────────────────────────────────────────────────────────────────
@app.get("/api/products")
async def get_products():
    df = read_sheet('Products')
    if df.empty:
        return []
    return df.to_dict(orient='records')

@app.post("/api/products")
async def create_product(product: ProductModel):
    product.id = gen_id()
    df = read_sheet('Products')

    # Validation: cost_price <= mrp
    if product.cost_price > product.mrp:
        raise HTTPException(status_code=400, detail=f"Cost price (₹{product.cost_price}) cannot exceed MRP (₹{product.mrp})")

    if not df.empty and (df['name'] == product.name).any():
        raise HTTPException(status_code=400, detail="Product name already exists")

    new_row = pd.DataFrame([product.dict()])
    df = pd.concat([df, new_row], ignore_index=True)
    write_sheet(df, 'Products')
    return {"id": product.id, "message": "Product created"}

@app.put("/api/products/{product_id}")
async def update_product(product_id: str, product: ProductModel):
    df = read_sheet('Products')

    if not (df['id'] == product_id).any():
        raise HTTPException(status_code=404, detail="Product not found")

    if product.cost_price > product.mrp:
        raise HTTPException(status_code=400, detail=f"Cost price cannot exceed MRP")

    df.loc[df['id'] == product_id] = product.dict()
    write_sheet(df, 'Products')
    return {"message": "Product updated"}

@app.delete("/api/products/{product_id}")
async def delete_product(product_id: str):
    df = read_sheet('Products')
    df = df[df['id'] != product_id]
    write_sheet(df, 'Products')
    return {"message": "Product deleted"}

# ─────────────────────────────────────────────────────────────────
# PURCHASES
# ─────────────────────────────────────────────────────────────────
@app.get("/api/purchases")
async def get_purchases():
    df = read_sheet('Purchases')
    if df.empty:
        return []

    products_df = read_sheet('Products')
    vendors_df = read_sheet('Vendors')

    if not products_df.empty:
        df = df.merge(products_df[['id', 'name', 'cost_price', 'mrp', 'gst_rate']],
                     left_on='product_id', right_on='id', how='left', suffixes=('', '_prod'))

    if not vendors_df.empty:
        df = df.merge(vendors_df[['id', 'name']],
                     left_on='vendor_id', right_on='id', how='left', suffixes=('', '_vendor'))

    return df.to_dict(orient='records')

@app.post("/api/purchases")
async def create_purchase(purchase: PurchaseModel):
    purchase.id = gen_id()
    df = read_sheet('Purchases')

    if not df.empty and (df['invoice_no'] == purchase.invoice_no).any():
        raise HTTPException(status_code=400, detail="Invoice number already exists")

    new_row = pd.DataFrame([purchase.dict()])
    df = pd.concat([df, new_row], ignore_index=True)
    write_sheet(df, 'Purchases')
    return {"id": purchase.id, "message": "Purchase recorded"}

@app.delete("/api/purchases/{purchase_id}")
async def delete_purchase(purchase_id: str):
    df = read_sheet('Purchases')
    df = df[df['id'] != purchase_id]
    write_sheet(df, 'Purchases')
    return {"message": "Purchase deleted"}

# ─────────────────────────────────────────────────────────────────
# SALES - PRICE VALIDATION: cost_price <= selling_price <= mrp
# ─────────────────────────────────────────────────────────────────
@app.get("/api/sales")
async def get_sales():
    df = read_sheet('Sales')
    if df.empty:
        return []

    products_df = read_sheet('Products')
    if not products_df.empty:
        df = df.merge(products_df[['id', 'name', 'cost_price', 'mrp', 'gst_rate']],
                     left_on='product_id', right_on='id', how='left')

    return df.to_dict(orient='records')

@app.post("/api/sales")
async def create_sale(sale: SalesModel):
    sale.id = gen_id()

    products_df = read_sheet('Products')
    product = products_df[products_df['id'] == sale.product_id]

    if product.empty:
        raise HTTPException(status_code=404, detail="Product not found")

    cost_price = float(product.iloc[0]['cost_price'])
    mrp = float(product.iloc[0]['mrp'])

    # PRICE VALIDATION
    if sale.selling_price < cost_price:
        raise HTTPException(status_code=400, detail=f"❌ Cannot sell below COST PRICE ₹{cost_price}")

    if sale.selling_price > mrp:
        raise HTTPException(status_code=400, detail=f"❌ Cannot sell above MRP ₹{mrp}")

    # STOCK CHECK
    purchases_df = read_sheet('Purchases')
    sales_df = read_sheet('Sales')

    total_purchased = purchases_df[purchases_df['product_id'] == sale.product_id]['quantity'].sum() if not purchases_df.empty else 0
    total_sold = sales_df[sales_df['product_id'] == sale.product_id]['quantity'].sum() if not sales_df.empty else 0
    available = total_purchased - total_sold

    if sale.quantity > available:
        raise HTTPException(status_code=400, detail=f"❌ Stock insufficient. Available: {available}")

    df = read_sheet('Sales')
    new_row = pd.DataFrame([sale.dict()])
    df = pd.concat([df, new_row], ignore_index=True)
    write_sheet(df, 'Sales')

    return {"id": sale.id, "message": "Sale recorded"}

@app.delete("/api/sales/{sale_id}")
async def delete_sale(sale_id: str):
    df = read_sheet('Sales')
    df = df[df['id'] != sale_id]
    write_sheet(df, 'Sales')
    return {"message": "Sale deleted"}

# ─────────────────────────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────────────────────────
@app.get("/api/dashboard")
async def get_dashboard():
    try:
        vendors_df = read_sheet('Vendors')
        products_df = read_sheet('Products')
        purchases_df = read_sheet('Purchases')
        sales_df = read_sheet('Sales')

        total_vendors = len(vendors_df) if not vendors_df.empty else 0
        total_products = len(products_df) if not products_df.empty else 0
        total_purchases = purchases_df['amount_paid'].sum() if not purchases_df.empty else 0
        total_sales = (sales_df['selling_price'] * sales_df['quantity']).sum() if not sales_df.empty else 0

        total_stock = 0
        if not purchases_df.empty:
            total_stock = purchases_df['quantity'].sum()
            if not sales_df.empty:
                total_stock -= sales_df['quantity'].sum()

        expiring_soon = 0
        if not purchases_df.empty:
            from datetime import timedelta
            today = datetime.now().date()
            three_months = today + timedelta(days=90)
            purchases_df['expiry_date'] = pd.to_datetime(purchases_df['expiry_date']).dt.date
            expiring_soon = len(purchases_df[(purchases_df['expiry_date'] <= three_months) & (purchases_df['expiry_date'] >= today)])

        return {
            "total_vendors": total_vendors,
            "total_products": total_products,
            "total_purchases": round(total_purchases, 2),
            "total_sales": round(total_sales, 2),
            "total_stock": int(total_stock),
            "expiring_soon": expiring_soon
        }
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/stock")
async def get_stock():
    try:
        products_df = read_sheet('Products')
        purchases_df = read_sheet('Purchases')
        sales_df = read_sheet('Sales')

        stock_data = []
        if not products_df.empty:
            for _, prod in products_df.iterrows():
                purchased = purchases_df[purchases_df['product_id'] == prod['id']]['quantity'].sum() if not purchases_df.empty else 0
                sold = sales_df[sales_df['product_id'] == prod['id']]['quantity'].sum() if not sales_df.empty else 0
                available = purchased - sold

                stock_data.append({
                    "product_id": prod['id'],
                    "product_name": prod['name'],
                    "cost_price": prod['cost_price'],
                    "mrp": prod['mrp'],
                    "purchased": int(purchased),
                    "sold": int(sold),
                    "available": int(available)
                })

        return stock_data
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/expiry")
async def get_expiry():
    try:
        purchases_df = read_sheet('Purchases')
        products_df = read_sheet('Products')

        if purchases_df.empty:
            return []

        purchases_df['expiry_date'] = pd.to_datetime(purchases_df['expiry_date'])
        if not products_df.empty:
            purchases_df = purchases_df.merge(products_df[['id', 'name']],
                                             left_on='product_id', right_on='id', how='left')

        from datetime import timedelta
        today = datetime.now().date()

        expiry_data = []
        for _, row in purchases_df.iterrows():
            expiry = row['expiry_date'].date()
            if expiry < today:
                status = "expired"
            elif expiry < today + timedelta(days=30):
                status = "expiring_soon"
            elif expiry < today + timedelta(days=90):
                status = "expiring_3m"
            else:
                status = "good"

            expiry_data.append({
                "product_id": row['product_id'],
                "product_name": row['name'],
                "batch_no": row['batch_no'],
                "quantity": row['quantity'],
                "expiry_date": str(row['expiry_date'].date()),
                "status": status
            })

        return expiry_data
    except Exception as e:
        return {"error": str(e)}


@app.get("/")
async def root():
    """Serve the frontend HTML"""
    html_file = Path(__file__).parent.parent / "graftcare.html"
    if html_file.exists():
        return FileResponse(html_file, media_type="text/html")
    return {"message": "Graft Care Pharma API"}

@app.get("/api/health")
async def health():
    return {"status": "running"}
